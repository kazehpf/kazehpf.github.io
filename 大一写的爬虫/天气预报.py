import requests
from bs4 import BeautifulSoup
import openpyxl
r=requests.get('https://weather.cma.cn/web/weather/map.html')
t=r.text
r.encoding=r.apparent_encoding
s=BeautifulSoup(t,features="html.parser")
t1=s.findAll('div',attrs={'class':'hp'})
city=t1[1].findAll('div',attrs={'class':"col-xs-2 city"})
wb=openpyxl.Workbook()
sheet=wb.create_sheet(title='sheet2')

# for row in sheet.rows:
#     for cell in row:
#       cell.value=None
sheet['A1']='城市'
sheet['B1']='天气'
sheet['C1']='风力'
sheet['D1']='最高温度'
sheet['E1']='最低温度'
wb.save('省会城市实时天气预报.xlsx')
for i in range(34):
    wb = openpyxl.load_workbook('省会城市实时天气预报.xlsx')
    sheet=wb['sheet2']
    t2 = city[i].find('a').get('href')
    rr=requests.get('https://weather.cma.cn'+t2)
    ss=BeautifulSoup(rr.text,features="html.parser")
    value=ss.find('div',attrs={'class':'hp'}).findAll('div',attrs={'class':'day-item'})
    ls=[city[i].string,value[2].string,value[3].string+value[4].string,value[5].find('div',attrs={'class':'high'}).string,value[5].find('div',attrs={'class':'low'}).string]
    sheet.append(ls)
    wb.save('省会城市实时天气预报.xlsx')
    print('已完成{}个城市的写入'.format(i+1))
print('全国城市当天天气已经全部录入')